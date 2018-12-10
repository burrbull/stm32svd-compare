#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Feb  3 10:33:35 2018

@author: burrbull
"""
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook

syns = {
		"ADC_Common" : "ADC",
		"DAC2" : "DAC",
		"Ethernet" : "ETHERNET",
		"Flash" : "FLASH",
		"SAI1" : "SAI",
		"TIMs" : "TIM",
		"USB_OTG" : "USB_OTG_FS",
		"USB_FS" : "USB",
		"DSIHOST" : "DSI",
		}
def gr_name(name):
	return syns[name] if name in syns else name


pth = os.getcwd()+"/"

from devicedict import devices

from hashlib import blake2b

def get_registers(p):
	string = ET.tostring(p.find("registers"),encoding="utf-8")
	h = blake2b(digest_size=10)
	h.update(string)
	return h.hexdigest(), string
	#return hash(string), string

def compare_svd(grname, same=True, excelOnly=False):
	pdict = {}
	for fname in devices[grname]:
		etree = ET.parse(pth+fname+".svd")
		
		a = etree.find("peripherals")
		
		for p in a.iter("peripheral"):
			nm = p.find("name")
			printed = "{} {}".format(fname, nm.text)
			short_name = fname.replace("STM32", "")
			newfname = "{}.{}".format(nm.text, short_name)
			sameperiphs = ""
			if 'derivedFrom' in p.attrib:
				printed = "- "+printed
			else:
				groupName = gr_name(p.find("groupName").text)
				fpath = pth+grname+"/"+groupName
				if groupName not in pdict: # создать группу (ADC, SPI и т.д.)
					pdict[groupName] = {}
				 # вырезать регистры
				hsh, string = get_registers(p)
				if hsh in pdict[groupName]:
					pdict[groupName][hsh].append(newfname)
					printed = "  "+printed
				else:
					printed = "* "+printed
					pdict[groupName][hsh] = [newfname]
					groupName = gr_name(p.find("groupName").text)
					rt = ET.ElementTree(p.find("registers"))
					if not excelOnly:
						if not os.path.exists(fpath):
							os.makedirs(fpath)
						rt.write("{}/{}_{}.xml".format(fpath, groupName, hsh))
				sameperiphs += "{:20} -> {}.{}\n".format(str(hsh), short_name, nm.text)
				p.remove(p.find("registers"))
				h = ET.Element("hash")
				h.text = str(hsh)
				p.append(h)
			if same and (not excelOnly):
				if not os.path.exists(fpath):
					os.makedirs(fpath)
				with open("{}/sameperiphs.txt".format(fpath), "a") as f:
					f.write(sameperiphs)
			print(printed)
		if not excelOnly:
			etree.write("{}/{}.xml".format(pth+grname,fname))
	
	import pandas as pd
	
	d = [[per, str(h) , ", ".join(v)] for per in pdict for h, v in pdict[per].items()]
	#print(d)
	df = pd.DataFrame(d)
	excelname = pth+grname+'.xlsx'
	df.to_excel(excelname)

def compare_svds(ids, same=True, excelOnly=False):
	wb = Workbook()
	for k in ids:
		compare_svd(k, same, excelOnly)
		rows = reader(pth+k+'.xlsx')
		sheet = wb.create_sheet(k)
		for row in rows:
			sheet.append(row)
		os.remove(pth+k+'.xlsx')
		if same and (not excelOnly):
			for dirpath, _, filenames in os.walk(pth+k+'/'):
				if "sameperiphs.txt" in filenames:
					os.system('sort -o {0} {0}'.format(dirpath+"/sameperiphs.txt"))
	wb.save(pth+"Peripherals.xlsx")
	
	
def reader(file):
	wb_sheet = load_workbook(file).active
	rows = []
	for row in wb_sheet.iter_rows():
		row_data = []
		for cell in row:
			row_data.append(cell.value)
			rows.append(row_data)
	return rows

def restore_svd(fpath, infile, outpath, patch=True):
	print(infile)
	etree = ET.parse(fpath+infile+'.xml')
	a = etree.find("peripherals")
	for p in a.iter("peripheral"):
		if 'derivedFrom' in p.attrib:
			pass
		else:
			groupName = gr_name(p.find("groupName").text)
			h = p.find("hash")
			hsh = h.text
			print("  ", groupName, hsh)
			p.remove(h)
			rtree = ET.parse("{0}{1}/{1}_{2}.xml".format(fpath, groupName, hsh))
			r = rtree.getroot()
			p.append(r)
	tmp = outpath+infile+".svd"
	etree.write(tmp)
	os.system('xmllint --format "{0}{1}.svd" --output "{0}{1}.svd"'.format(outpath, infile))
	if patch:
		os.system('patch "{0}{1}.svd" "{0}patches/{1}.patch"'.format(outpath, infile))
	os.system('unix2dos "{0}{1}.svd"'.format(outpath, infile))

def restore_svds(gr, patch=True):
	for d in devices[gr]:
		restore_svd(pth+gr+"/", d, pth, patch)

def diff(fpath, groupName, ls=[]):
	import itertools
	fpath = fpath+groupName+"/"
	lst = [f for f in os.listdir(fpath) if f.endswith(".xml")]
	s = ""
	for f1, f2 in itertools.combinations(lst, 2):
		s += ("\n#############################"
			"### {} -> {} ###\n\n".format(f1, f2)
			)
		t = os.popen("diff -u3 {0}{1} {0}{2}".format(fpath, f1, f2))
		ss = t.read()
		s += ss
		s += "\n\n\n"
		ls.append([groupName, f1, f2, len(ss)])
	if len(lst) > 1:
		with open("{}{}_diff.txt", "w".format(fpath, groupName)) as f:
			f.write(s)

def diff_svds(folder):
	fpath = pth+folder+"/"
	lst = []
	for gr in os.listdir(fpath):
		if os.path.isdir(fpath+gr):
			diff(fpath, gr, lst)
	lst = sorted(lst, key = lambda r: (r[0], r[3]))
	for row in lst:
		s = " meld {0}/{1} {0}/{2}".format(row[0], row[1], row[2])
		row.append(s)
	import pandas as pd
	df = pd.DataFrame(lst)
	excelname = pth+"Diffs.xlsx"
	df.to_excel(excelname)

def sort_svd(grname):
	fpath = pth+"sorted/"
		
	for fname in devices[grname]:
		sort_single(pth, fpath, fname)
		
def sort_single(inpath, outpath, fname, down=True):
	if not os.path.exists(outpath):
		os.makedirs(outpath)
	print(fname)
	etree = ET.parse(inpath+fname+".svd")
	
	pp = etree.find("peripherals")

	for p in pp.iter("peripheral"):
		print("\t"+p.findtext("name"))
		if 'derivedFrom' in p.attrib:
			continue
		rr = p.find("registers")
		sort_registers(rr)
		
		for r in rr.iter("register"):
			ff = r.find("fields")
			sort_fields(ff, down)
		
		for c in rr.iter("cluster"):
			#sort_registers(c)
			for r in c.iter("register"):
				ff = r.find("fields")
				sort_fields(ff, down)
	etree.write(outpath+fname+".svd", encoding="utf-8")
	os.system('xmllint --format "{0}{1}.svd" --output "{0}{1}.svd"'.format(outpath, fname))
	os.system('patch "{0}{1}.svd" "{0}/../patches/{1}.patch"'.format(outpath, fname))
	os.system('unix2dos "{0}{1}.svd"'.format(outpath, fname))


def sort_fields(ff, down=True):
	if ff:
		fields = []
		for f in ff:
			key = int(f.findtext("bitOffset"))
			fields.append((key, f))
		#print(fields)
		if down:
			fields.sort(reverse=True, key=lambda x: x[0])
		else:
			fields.sort(key=lambda x: x[0])
		ff[:] = [item[-1] for item in fields]
		
def sort_registers(rr):
	if rr:
		registers = []
		for r in rr:
		#	print(r.find("name").text)
			txt = r.findtext("addressOffset")
			#print(txt)
			key = int(txt, 16)
			registers.append((key, r))
		registers.sort(key=lambda x: x[0])
		rr[:] = [item[-1] for item in registers]

def make_patches(gr):
	for d in devices[gr]:
		forig = pth+"orig/"+d+".svd"
		fnew = pth+d+".svd"
		fpatch = pth+"patches/"+d+".patch"
		s = 'diff -u3 "{}" "{}"'.format(fnew, forig)
		t = os.popen(s)
		st = t.read()
		with open(fpatch, "w") as f:
			f.write(st)


def svd_statistics(grname):
	for fname in devices[grname]:
		etree = ET.parse(pth+fname+".svd")
		
		a = etree.find("peripherals")
		print(fname, len(a.findall("peripheral")))
		#for p in a.iter("peripheral"):
		#	print(p)


def copy_enums(fromfile, tofile, outfile):
	fromtree = ET.parse(fromfile)
	outtree = ET.parse(tofile)
	
	pp = fromtree.find("peripherals")
	ppout = outtree.find("peripherals")

	for p in pp.iter("peripheral"):
		pname = p.findtext("name")
		if 'derivedFrom' in p.attrib:
			continue
		ppath = SvdPath.from_str(pname)
		pout = ppout.find("peripheral[name='{}']".format(pname))
		if not pout:
			print("Peripheral {} not found".format(pname))
			continue
		
		rr = p.find("registers")
		rrout = pout.find("registers")
		copy_register_enums(rr, rrout, ppath)
		
		for c in rr.iter("cluster"):
			cname = c.findtext("name")
			cout = rrout.find("cluster[name='{}']".format(cname))
			cpath = ppath+cname
			if not cout:
				print("Cluster {} not found".format(cpath))
				continue
			copy_register_enums(c, cout, cpath)
	outtree.write(outfile, encoding="utf-8")
	os.system('xmllint --format "{0}" --output "{0}"'.format(outfile))
	os.system('unix2dos "{}"'.format(outfile))

def copy_register_enums(rr, rrout, rrpath):
	for r in rr.findall("register"):
		rname = r.findtext("name")
		rpath = rrpath + rname
		rout = rrout.find("register[name='{}']".format(rname))
		if not rout:
			print("Register {} not found".format(rpath))
			continue
		ff = r.find("fields")
		ffout = rout.find("fields")
		for f in ff.iter("field"):
			fname = f.findtext("name")
			fpath = rpath + fname
			fout = ffout.find("field[name='{}']".format(fname))
			if not fout:
				print("Field {} not found".format(fpath))
				continue
			for ev in f.iter("enumeratedValues"):
				fout.append(ev)
			

class SvdPath(list):
	def from_str(s):
		return SvdPath(s.split("."))
	def __init__(self, *args, **kwargs):
		super().__init__(*args, **kwargs)
	def __add__(self, other):
		return SvdPath(list(self)+[other])
	def __str__(self):
		return ".".join(self)
		